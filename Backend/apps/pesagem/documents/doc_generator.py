

from datetime import date, datetime
from decimal import Decimal
from typing import Any, Optional
from io import BytesIO

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from apps.pesagem.utils.excel_utils import (
    aplicar_estilo_completo,
    formatar_coluna_data,
    formatar_coluna_moeda,
    formatar_coluna_decimal,
)


class PesagemExcelConfig:
    
    def __init__(
        self,
        titulo: str = "Relatório de Pesagens",
        com_filtros: bool = True,
        com_zebra: bool = True,
        congelar_cabecalho: bool = True,
        colunas_data: Optional[list[str]] = None,
        colunas_moeda: Optional[list[str]] = None,
        colunas_decimal: Optional[list[str]] = None,
    ):
        self.titulo = titulo
        self.com_filtros = com_filtros
        self.com_zebra = com_zebra
        self.congelar_cabecalho = congelar_cabecalho
        self.colunas_data = colunas_data or []
        self.colunas_moeda = colunas_moeda or []
        self.colunas_decimal = colunas_decimal or []


class PesagemExcelDocument:
    
    def __init__(self, config: Optional[PesagemExcelConfig] = None):
        self.config = config or PesagemExcelConfig()
        self.wb = Workbook()
        self.ws: Worksheet = self.wb.active
        self.ws.title = self.config.titulo[:31]  
        self._headers_map: dict[str, str] = {}
    
    @staticmethod
    def gerar(
        dados: list[dict],
        config: Optional[PesagemExcelConfig] = None
    ) -> Workbook:
        doc = PesagemExcelDocument(config)
        return doc._gerar_documento(dados)
    
    @staticmethod
    def gerar_buffer(
        dados: list[dict],
        config: Optional[PesagemExcelConfig] = None
    ) -> BytesIO:
        wb = PesagemExcelDocument.gerar(dados, config)
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def _gerar_documento(self, dados: list[dict]) -> Workbook:
        if not dados:
            self.ws.append(["Nenhum dado encontrado"])
            aplicar_estilo_completo(
                self.ws,
                com_filtros=False,
                com_zebra=False,
                congelar=False
            )
            return self.wb
        headers = list(dados[0].keys())
        headers_formatados = self._formatar_headers(headers)
        self.ws.append(headers_formatados)
        self._criar_mapa_colunas(headers)
        for row in dados:
            linha_formatada = self._formatar_linha(row.values())
            self.ws.append(linha_formatada)
        self._aplicar_formatacao_colunas()
        aplicar_estilo_completo(
            self.ws,
            com_filtros=self.config.com_filtros,
            com_zebra=self.config.com_zebra,
            congelar=self.config.congelar_cabecalho
        )
        
        return self.wb
    
    def _criar_mapa_colunas(self, headers: list[str]) -> None:
        from openpyxl.utils import get_column_letter
        
        for idx, header in enumerate(headers, start=1):
            self._headers_map[header] = get_column_letter(idx)
    
    def _aplicar_formatacao_colunas(self) -> None:

        for coluna in self.config.colunas_data:
            if coluna in self._headers_map:
                formatar_coluna_data(self.ws, self._headers_map[coluna])
        
        for coluna in self.config.colunas_moeda:
            if coluna in self._headers_map:
                formatar_coluna_moeda(self.ws, self._headers_map[coluna])
        
        for coluna in self.config.colunas_decimal:
            if coluna in self._headers_map:
                formatar_coluna_decimal(self.ws, self._headers_map[coluna])
    
    @staticmethod
    def _formatar_headers(headers: list[str]) -> list[str]:
        return [
            h.replace("_", " ").title()
            for h in headers
        ]
    
    @staticmethod
    def _formatar_linha(valores) -> list[Any]:
        linha = []
        for v in valores:
            if isinstance(v, (date, datetime)):
                linha.append(v)
            elif isinstance(v, Decimal):
                linha.append(float(v))
            elif v is None:
                linha.append("")
            else:
                linha.append(v)
        return linha
