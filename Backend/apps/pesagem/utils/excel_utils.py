from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter




class Cores:
   
    HEADER_BG = "1F4E79"          # Azul escuro profissional
    HEADER_FONT = "FFFFFF"        # Branco
    ZEBRA_LIGHT = "F2F2F2"        # Cinza claro
    ZEBRA_DARK = "FFFFFF"         # Branco
    BORDER = "BFBFBF"             # Cinza médio
    SUCCESS = "27AE60"            # Verde
    WARNING = "F39C12"            # Amarelo
    DANGER = "E74C3C"             # Vermelho



HEADER_FILL = PatternFill(
    start_color=Cores.HEADER_BG,
    end_color=Cores.HEADER_BG,
    fill_type="solid"
)

HEADER_FONT = Font(
    bold=True,
    color=Cores.HEADER_FONT,
    size=11,
    name="Calibri"
)

BODY_FONT = Font(
    size=10,
    name="Calibri"
)

CENTER_ALIGNMENT = Alignment(
    horizontal="center",
    vertical="center",
    wrap_text=False
)

LEFT_ALIGNMENT = Alignment(
    horizontal="left",
    vertical="center",
    wrap_text=False
)

THIN_BORDER = Border(
    left=Side(style="thin", color=Cores.BORDER),
    right=Side(style="thin", color=Cores.BORDER),
    top=Side(style="thin", color=Cores.BORDER),
    bottom=Side(style="thin", color=Cores.BORDER),
)

ZEBRA_FILL_LIGHT = PatternFill(
    start_color=Cores.ZEBRA_LIGHT,
    end_color=Cores.ZEBRA_LIGHT,
    fill_type="solid"
)

ZEBRA_FILL_DARK = PatternFill(
    start_color=Cores.ZEBRA_DARK,
    end_color=Cores.ZEBRA_DARK,
    fill_type="solid"
)



def aplicar_estilo_cabecalho(ws: Worksheet, altura_linha: int = 25) -> None:
    ws.row_dimensions[1].height = altura_linha
    
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGNMENT
        cell.border = THIN_BORDER


def aplicar_bordas(ws: Worksheet) -> None:
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.border = THIN_BORDER


def aplicar_zebra(ws: Worksheet, linha_inicio: int = 2) -> None:
    for idx, row in enumerate(ws.iter_rows(min_row=linha_inicio), start=linha_inicio):
        fill = ZEBRA_FILL_LIGHT if idx % 2 == 0 else ZEBRA_FILL_DARK
        for cell in row:
            if cell.value is not None:
                cell.fill = fill
                cell.font = BODY_FONT
                cell.alignment = LEFT_ALIGNMENT


def ajustar_largura_colunas(ws: Worksheet, padding: int = 3, min_width: int = 10, max_width: int = 50) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            if cell.value:
                cell_length = len(str(cell.value))
                max_length = max(max_length, cell_length)

        adjusted_width = min(max(max_length + padding, min_width), max_width)
        ws.column_dimensions[column_letter].width = adjusted_width


def aplicar_filtros(ws: Worksheet) -> None:
    if ws.max_row < 1 or ws.max_column < 1:
        return
    
    ultima_coluna = get_column_letter(ws.max_column)
    filtro_range = f"A1:{ultima_coluna}{ws.max_row}"
    ws.auto_filter.ref = filtro_range


def congelar_cabecalho(ws: Worksheet) -> None:
    ws.freeze_panes = "A2"
def formatar_coluna_data(ws: Worksheet, coluna: str) -> None:

    for cell in ws[coluna][1:]:
        if cell.value:
            cell.number_format = "DD/MM/YYYY"
            cell.alignment = CENTER_ALIGNMENT


def formatar_coluna_moeda(ws: Worksheet, coluna: str) -> None:
    for cell in ws[coluna][1:]:
        if cell.value:
            cell.number_format = 'R$ #,##0.00'
            cell.alignment = CENTER_ALIGNMENT


def formatar_coluna_decimal(ws: Worksheet, coluna: str, casas: int = 2) -> None:
    formato = "#,##0." + "0" * casas
    for cell in ws[coluna][1:]:
        if cell.value:
            cell.number_format = formato
            cell.alignment = CENTER_ALIGNMENT


def formatar_coluna_percentual(ws: Worksheet, coluna: str) -> None:
    for cell in ws[coluna][1:]:
        if cell.value:
            cell.number_format = '0.00%'
            cell.alignment = CENTER_ALIGNMENT

def aplicar_estilo_completo(
    ws: Worksheet,
    com_filtros: bool = True,
    com_zebra: bool = True,
    congelar: bool = True
) -> None:
    aplicar_estilo_cabecalho(ws)
    aplicar_bordas(ws)
    
    if com_zebra:
        aplicar_zebra(ws)
    
    ajustar_largura_colunas(ws)
    
    if com_filtros:
        aplicar_filtros(ws)
    
    if congelar:
        congelar_cabecalho(ws)
