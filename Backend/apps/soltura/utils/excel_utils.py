# utils.py
import os
from datetime import datetime, timedelta
from collections import Counter, defaultdict
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from django.conf import settings

def determinar_turno(hora_saida):
    """
    Determina o turno baseado no horário de saída.
    Manhã: 06:00 - 13:59
    Tarde: 14:00 - 21:59
    Noite: 22:00 - 05:59
    """
    if not hora_saida:
        return "Sem Horário"
    
    hora = hora_saida.hour
    
    if 6 <= hora < 14:
        return "Manhã"
    elif 14 <= hora < 22:
        return "Tarde"
    else:
        return "Noite"

def gerar_grafico_turnos_por_mes(dados_por_dia, data_inicio, data_final, caminho_base):
    """
    Cria um gráfico separado para cada mês no range de datas
    Gera gráficos por mês, um para cada mês no período selecionado.
    Se apenas 1 mês foi selecionado, gera apenas um gráfico.
    """
    if not dados_por_dia:
        return []
    
    try:
        # Organizar dados por mês
        dados_por_mes = defaultdict(dict)
        for data, turnos in dados_por_dia.items():
            mes_ano = (data.year, data.month)
            dados_por_mes[mes_ano][data] = turnos
        
        # Se apenas 1 mês, retorna gráfico único
        if len(dados_por_mes) == 1:
            mes_ano = list(dados_por_mes.keys())[0]
            caminho = caminho_base.replace('_grafico.png', f'_grafico_unico.png')
            return [gerar_grafico_mes_individual(dados_por_mes[mes_ano], mes_ano, caminho)]
        
        # Se múltiplos meses, cria um gráfico para cada
        graficos_gerados = []
        for mes_ano in sorted(dados_por_mes.keys()):
            caminho = caminho_base.replace('_grafico.png', f'_grafico_{mes_ano[0]}_{mes_ano[1]:02d}.png')
            resultado = gerar_grafico_mes_individual(dados_por_mes[mes_ano], mes_ano, caminho)
            if resultado:
                graficos_gerados.append(resultado)
        
        return graficos_gerados
    
    except Exception as e:
        print(f"Erro ao gerar gráficos por mês: {e}")
        return []

def gerar_grafico_comparacao_meses(dados_por_dia, caminho_grafico):
    """
    Gera um gráfico de comparação entre meses mostrando aumento/decremento por turno
    Retorna o caminho do gráfico e uma lista de análises de variação.
    """
    if not dados_por_dia:
        return None, []
    
    try:
        meses_pt = {
            1: 'Jan', 2: 'Fev', 3: 'Mar', 4: 'Abr',
            5: 'Mai', 6: 'Jun', 7: 'Jul', 8: 'Ago',
            9: 'Set', 10: 'Out', 11: 'Nov', 12: 'Dez'
        }
        
        # Organizar dados por mês e turno
        dados_por_mes = defaultdict(lambda: {"Manhã": 0, "Tarde": 0, "Noite": 0})
        for data, turnos in dados_por_dia.items():
            mes_ano = (data.year, data.month)
            for turno, qtd in turnos.items():
                dados_por_mes[mes_ano][turno] += qtd
        
        # Se menos de 2 meses, não gera comparação
        if len(dados_por_mes) < 2:
            return None, []
        
        meses_ordenados = sorted(dados_por_mes.keys())
        labels = [f"{meses_pt[mes[1]]}/{mes[0]}" for mes in meses_ordenados]
        
        turnos = ["Manhã", "Tarde", "Noite"]
        cores = {
            "Manhã": "#F4C430",
            "Tarde": "#FF9500",
            "Noite": "#4A90E2"
        }
        
        # Preparar dados por turno
        dados_turnos = {turno: [] for turno in turnos}
        for mes_ano in meses_ordenados:
            for turno in turnos:
                dados_turnos[turno].append(dados_por_mes[mes_ano][turno])
        
        analises = []
        for i in range(1, len(meses_ordenados)):
            mes_atual = meses_ordenados[i]
            mes_anterior = meses_ordenados[i-1]
            
            for turno in turnos:
                valor_atual = dados_por_mes[mes_atual][turno]
                valor_anterior = dados_por_mes[mes_anterior][turno]
                
                if valor_anterior > 0:
                    variacao = ((valor_atual - valor_anterior) / valor_anterior) * 100
                    if abs(variacao) > 5:  # Somente variações significativas (>5%)
                        tipo = "crescimento" if variacao > 0 else "decrescimento"
                        analises.append(
                            f"• {turno}: {tipo} de {abs(variacao):.1f}% "
                            f"({mes_anterior[1]}/{mes_anterior[0]} → {mes_atual[1]}/{mes_atual[0]})"
                        )
        
        # Criar gráfico
        fig, ax = plt.subplots(figsize=(16, 8), dpi=120, facecolor='white')
        ax.set_facecolor('#FAFAFA')
        
        x = np.arange(len(labels))
        width = 0.25
        
        # Plotar barras agrupadas
        for idx, turno in enumerate(turnos):
            offset = (idx - 1) * width
            bars = ax.bar(x + offset, dados_turnos[turno], width,
                         label=turno, color=cores[turno], alpha=0.9,
                         edgecolor='white', linewidth=1)
            
            # Adicionar valores nas barras
            for bar in bars:
                height = bar.get_height()
                if height > 0:
                    ax.text(bar.get_x() + bar.get_width()/2., height + 0.5,
                           f'{int(height)}',
                           ha='center', va='bottom', fontsize=10,
                           fontweight='bold', color='#333333')
        
        ax.set_title('Comparação Mensal de Saídas por Turno', 
                    fontsize=16, fontweight='bold', pad=20, color='#333333')
        ax.set_xlabel('Mês', fontsize=13, fontweight='bold', color='#333333', labelpad=10)
        ax.set_ylabel('Número de Saídas', fontsize=13, fontweight='bold', color='#333333')
        
        ax.set_xticks(x)
        ax.set_xticklabels(labels, fontsize=11)
        ax.tick_params(axis='y', labelsize=11, colors='#333333')
        
        # Grid
        ax.grid(True, alpha=0.25, linestyle='--', linewidth=0.7, color='#CCCCCC', axis='y')
        ax.set_axisbelow(True)
        
        # Legenda
        ax.legend(loc='upper center', bbox_to_anchor=(0.5, -0.08), fontsize=13,
                 framealpha=0.95, edgecolor='#E0E0E0', ncol=3)
        
        # Limite Y
        max_value = max([max(dados_turnos[turno]) for turno in turnos if dados_turnos[turno]])
        ax.set_ylim(0, max_value * 1.15)
        
        plt.tight_layout()
        
        os.makedirs(os.path.dirname(caminho_grafico), exist_ok=True)
        plt.savefig(caminho_grafico, dpi=150, bbox_inches='tight',
                   facecolor='white', edgecolor='none', pad_inches=0.4)
        plt.close()
        
        return caminho_grafico, analises
    except Exception as e:
        plt.close('all')
        print(f"Erro ao gerar gráfico de comparação: {e}")
        return None, []

def gerar_grafico_mes_individual(dados_dias_mes, mes_ano, caminho_grafico):
    """
    Gera um gráfico individual para um mês específico
    """
    if not dados_dias_mes:
        return None
    
    try:
        meses_pt = {
            1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril',
            5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto',
            9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
        }
        
        datas = sorted(dados_dias_mes.keys())
        turnos = ["Manhã", "Tarde", "Noite"]
        cores = {
            "Manhã": "#F4C430",    
            "Tarde": "#FF9500",    
            "Noite": "#4A90E2"     
        }
        
        # Organizar dados por turno
        dados_turnos = {turno: [] for turno in turnos}
        for data in datas:
            for turno in turnos:
                dados_turnos[turno].append(dados_dias_mes[data].get(turno, 0))
        
        num_dias = len(datas)
        
        # Configurações dinâmicas baseadas no número de dias
        if num_dias <= 15:
            fig_width = 16
            width = 0.25
            rotation = 45
            fontsize = 9
        elif num_dias <= 31:
            fig_width = 20
            width = 0.23
            rotation = 45
            fontsize = 8
        else:
            fig_width = 24
            width = 0.22
            rotation = 45
            fontsize = 7
        
        fig, ax = plt.subplots(figsize=(fig_width, 8), dpi=120, facecolor='white')
        ax.set_facecolor('#FAFAFA')
        
        x = np.arange(len(datas))
        
        # Plotar barras agrupadas
        for idx, turno in enumerate(turnos):
            offset = (idx - 1) * width
            bars = ax.bar(x + offset, dados_turnos[turno], width,
                         label=turno, 
                         color=cores[turno],
                         alpha=0.9,
                         edgecolor='white',
                         linewidth=1)
            
            # Adicionar valores nas barras apenas se não estiver muito lotado
            if num_dias <= 31:
                for bar in bars:
                    height = bar.get_height()
                    if height > 0:
                        ax.text(bar.get_x() + bar.get_width()/2., height + 0.3,
                               f'{int(height)}',
                               ha='center', va='bottom', fontsize=fontsize, 
                               fontweight='bold', color='#333333')
        
        mes_nome = meses_pt[mes_ano[1]]
        ax.set_title(f'Análise de Saídas por Turno - {mes_nome}/{mes_ano[0]}', 
                    fontsize=16, fontweight='bold', pad=20, color='#333333')
        
        ax.set_xlabel('Data', fontsize=13, fontweight='bold', color='#333333', labelpad=10)
        ax.set_ylabel('Número de Saídas', fontsize=13, fontweight='bold', color='#333333')
        
        # Configurar todas as datas no eixo X
        datas_formatadas = [data.strftime('%d/%m') for data in datas]
        ax.set_xticks(x)
        ax.set_xticklabels(datas_formatadas, rotation=rotation, ha='right', fontsize=fontsize)
        
        ax.tick_params(axis='y', labelsize=11, colors='#333333')
        ax.tick_params(axis='x', which='major', length=5, width=1)
        
        # Grid
        ax.grid(True, alpha=0.25, linestyle='--', linewidth=0.7, color='#CCCCCC', axis='y')
        ax.set_axisbelow(True)
        
        # Legenda
        ax.legend(loc='upper center', bbox_to_anchor=(0.5, -0.12), fontsize=13, 
                 framealpha=0.95, edgecolor='#E0E0E0', fancybox=False, 
                 frameon=True, shadow=False, ncol=3)
        
        # Limite Y com margem
        max_value = max([max(dados_turnos[turno]) for turno in turnos if dados_turnos[turno]]) if any(dados_turnos.values()) else 1
        ax.set_ylim(0, max_value * 1.15)
        
        plt.tight_layout()
        
        os.makedirs(os.path.dirname(caminho_grafico), exist_ok=True)
        
        # Salvar com alta qualidade
        plt.savefig(caminho_grafico, dpi=150, bbox_inches='tight', 
                   facecolor='white', edgecolor='none', pad_inches=0.4)
        plt.close()
        
        return caminho_grafico
    except Exception as e:
        plt.close('all')
        print(f"Erro ao gerar gráfico do mês: {e}")
        return None

def limpar_arquivos_temporarios(caminho_base, grafico_comparacao=None):
    import glob
    for arquivo in glob.glob(caminho_base.replace('_grafico.png', '_grafico_*.png')):
        try:
            os.remove(arquivo)
        except Exception as e:
            print(f"Aviso: Não foi possível remover arquivo temporário: {e}")
    
    if grafico_comparacao and os.path.exists(grafico_comparacao):
        try:
            os.remove(grafico_comparacao)
        except Exception as e:
            print(f"Aviso: Não foi possível remover gráfico de comparação: {e}")

def get_styles():
    bold_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    summary_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    pa_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    return {
        'bold_font': bold_font,
        'title_font': title_font,
        'center_align': center_align,
        'left_align': left_align,
        'header_fill': header_fill,
        'summary_fill': summary_fill,
        'pa_fill': pa_fill,
        'thin_border': thin_border
    }



