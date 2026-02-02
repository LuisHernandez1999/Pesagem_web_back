# apps/soltura/documents/excel_generator.py
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from django.conf import settings
from apps.soltura.utils.excel_utils import (
    gerar_grafico_turnos_por_mes, gerar_grafico_comparacao_meses,
    limpar_arquivos_temporarios, get_styles
)
from apps.soltura.mappers.excel_mappers import get_header

class BaseReportGenerator:
    def __init__(
        self,
        service,
        processed_data,
        data_inicio: str,
        data_final: str,
        inicio_dt,
        final_dt,
    ):
        self.service = service
        self.processed_data = processed_data
        self.data_inicio = data_inicio
        self.data_final = data_final
        self.inicio_dt = inicio_dt
        self.final_dt = final_dt

        self.styles = get_styles()
        self.TOP_N = 40

    def generate(self, nome_arquivo: str = None):
        if not nome_arquivo:
            nome_arquivo = f"relatorio_{self.service.tipo_servico.lower()}_{self.data_inicio}_a_{self.data_final}.xlsx"

        caminho = os.path.join(settings.MEDIA_ROOT, nome_arquivo)
        os.makedirs(settings.MEDIA_ROOT, exist_ok=True)

        wb = Workbook()
        self._create_dados_sheet(wb)
        self._create_resumos_sheet(wb)
        self._create_analise_sheet(wb, nome_arquivo)

        wb.save(caminho)

        # Limpeza
        if self.processed_data['dados_por_dia_turno']:
            caminho_base = os.path.join(settings.MEDIA_ROOT, nome_arquivo.replace('.xlsx', '_grafico.png'))
            caminho_comparacao = caminho_base.replace('_grafico.png', '_comparacao.png')
            limpar_arquivos_temporarios(caminho_base, caminho_comparacao)

        return caminho

    def _create_dados_sheet(self, wb):
        ws_dados = wb.active
        ws_dados.title = f"Dados {self.service.tipo_servico}"

        # Cabeçalho principal
        ws_dados.merge_cells('A1:I1')
        ws_dados['A1'] = f"RELATÓRIO {self.service.tipo_servico.upper()}"
        ws_dados['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        ws_dados['A1'].alignment = self.styles['center_align']
        ws_dados['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        
        ws_dados.append([])
        ws_dados.append([f"Período: {self.data_inicio} a {self.data_final}"])
        ws_dados.cell(row=3, column=1).font = Font(bold=True, size=11)
        ws_dados.cell(row=3, column=1).alignment = self.styles['center_align']
        
        ws_dados.append([f"Total de registros: {self.processed_data['total_registros']}"])
        ws_dados.cell(row=4, column=1).font = Font(bold=True, size=11)
        ws_dados.cell(row=4, column=1).alignment = self.styles['center_align']
        
        ws_dados.append([])

        # Header de dados
        header = get_header(self.service.tipo_servico)
        ws_dados.append(header)
        header_row = ws_dados.max_row
        for col_num, value in enumerate(header, 1):
            cell = ws_dados.cell(row=header_row, column=col_num, value=value)
            cell.font = self.styles['bold_font']
            cell.alignment = self.styles['center_align']
            cell.fill = self.styles['header_fill']
            cell.border = self.styles['thin_border']

        # Dados
        for row_data in self.processed_data['rows']:
            ws_dados.append(row_data)
            current_row = ws_dados.max_row
            for col_num in range(1, len(row_data) + 1):
                cell = ws_dados.cell(row=current_row, column=col_num)
                cell.alignment = self.styles['center_align']
                cell.border = self.styles['thin_border']

        # Ajuste de colunas
        for col in ws_dados.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            adjusted_width = min(max_length + 3, 50)
            ws_dados.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

    def _create_resumos_sheet(self, wb):
        ws_resumos = wb.create_sheet("Resumos e Análises")
        ws_resumos.merge_cells('A1:I1')
        ws_resumos['A1'] = f"ANÁLISES E RESUMOS - {self.service.tipo_servico.upper()}"
        ws_resumos['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        ws_resumos['A1'].alignment = self.styles['center_align']
        ws_resumos['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        
        ws_resumos.append([])
        ws_resumos.append([f"Período: {self.data_inicio} a {self.data_final}"])
        ws_resumos.cell(row=3, column=1).font = Font(bold=True, size=11)
        ws_resumos.cell(row=3, column=1).alignment = self.styles['center_align']
        ws_resumos.append([])

        current_row = 5

        # Resumo Geral por Turno
        totais_turno = {"Manhã": 0, "Tarde": 0, "Noite": 0}
        for pa_data in self.processed_data['dados_por_pa'].values():
            for data_turnos in pa_data.values():
                for turno, qtd in data_turnos.items():
                    totais_turno[turno] += qtd
        total_geral_turnos = sum(totais_turno.values())

        ws_resumos.cell(row=current_row, column=1).value = "📊 RESUMO GERAL POR TURNO"
        ws_resumos.cell(row=current_row, column=1).font = self.styles['title_font']
        ws_resumos.cell(row=current_row, column=1).fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        current_row += 1

        # Headers
        ws_resumos.cell(row=current_row, column=1).value = "Turno"
        ws_resumos.cell(row=current_row, column=2).value = "Quantidade"
        ws_resumos.cell(row=current_row, column=3).value = "% do Total"
        for col in range(1, 4):
            cell = ws_resumos.cell(row=current_row, column=col)
            cell.font = self.styles['bold_font']
            cell.fill = self.styles['header_fill']
            cell.border = self.styles['thin_border']
            cell.alignment = self.styles['center_align']
        current_row += 1

        # Dados turnos
        for turno, emoji in [("Manhã", "🌅"), ("Tarde", "🌞"), ("Noite", "🌙")]:
            ws_resumos.cell(row=current_row, column=1).value = f"{emoji} {turno}"
            ws_resumos.cell(row=current_row, column=2).value = totais_turno[turno]
            ws_resumos.cell(row=current_row, column=3).value = f"{(totais_turno[turno]/total_geral_turnos*100):.1f}%" if total_geral_turnos > 0 else "0%"
            for col in range(1, 4):
                cell = ws_resumos.cell(row=current_row, column=col)
                cell.border = self.styles['thin_border']
                cell.alignment = self.styles['left_align'] if col == 1 else self.styles['center_align']
            current_row += 1

        ws_resumos.cell(row=current_row, column=1).value = "📈 Total"
        ws_resumos.cell(row=current_row, column=2).value = total_geral_turnos
        ws_resumos.cell(row=current_row, column=3).value = "100%"
        for col in range(1, 4):
            cell = ws_resumos.cell(row=current_row, column=col)
            cell.border = self.styles['thin_border']
            cell.font = self.styles['bold_font']
            cell.fill = PatternFill(start_color="FFF4CC", end_color="FFF4CC", fill_type="solid")
            cell.alignment = self.styles['left_align'] if col == 1 else self.styles['center_align']
        current_row += 2

        # Detalhamento por P.A e Data
        ws_resumos.cell(row=current_row, column=1).value = "📋 DETALHAMENTO POR P.A E DATA"
        ws_resumos.cell(row=current_row, column=1).font = self.styles['title_font']
        ws_resumos.cell(row=current_row, column=1).fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        current_row += 1

        ws_resumos.cell(row=current_row, column=1).value = "Data"
        ws_resumos.cell(row=current_row, column=2).value = "P.A (Garagem)"
        ws_resumos.cell(row=current_row, column=3).value = "Turno"
        ws_resumos.cell(row=current_row, column=4).value = "Quantidade de Saídas"
        for col in range(1, 5):
            cell = ws_resumos.cell(row=current_row, column=col)
            cell.font = self.styles['bold_font']
            cell.fill = self.styles['header_fill']
            cell.border = self.styles['thin_border']
            cell.alignment = self.styles['center_align']
        current_row += 1

        for pa in sorted(self.processed_data['dados_por_pa'].keys()):
            for data in sorted(self.processed_data['dados_por_pa'][pa].keys()):
                for turno in ["Manhã", "Tarde", "Noite"]:
                    qtd = self.processed_data['dados_por_pa'][pa][data].get(turno, 0)
                    if qtd > 0:
                        ws_resumos.cell(row=current_row, column=1).value = data.strftime('%d/%m/%Y')
                        ws_resumos.cell(row=current_row, column=2).value = pa
                        ws_resumos.cell(row=current_row, column=3).value = turno
                        ws_resumos.cell(row=current_row, column=4).value = qtd
                        for col in range(1, 5):
                            cell = ws_resumos.cell(row=current_row, column=col)
                            cell.border = self.styles['thin_border']
                            cell.alignment = self.styles['center_align']
                        current_row += 1
        current_row += 2

        # Resumo por Garagem
        self._add_resumo_section(ws_resumos, current_row, "RESUMO POR GARAGEM", self.processed_data['garagem_counter'], self.processed_data['total_registros'])
        current_row = ws_resumos.max_row + 1

        # Resumo por Líder
        self._add_resumo_section(ws_resumos, current_row, "RESUMO POR LÍDER", self.processed_data['lider_counter'], self.processed_data['total_registros'])
        current_row = ws_resumos.max_row + 1

        # Top Específico (Setor ou Rota)
        titulo_especifico = f"TOP {self.TOP_N} {self.service.campo_especifico.upper()}ES"
        self._add_top_section(ws_resumos, current_row, titulo_especifico, self.processed_data['especifico_counter'], self.processed_data['total_registros'])

        # Ajuste de colunas
        for col in ws_resumos.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            adjusted_width = min(max_length + 3, 60)
            ws_resumos.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

    def _add_resumo_section(self, ws, start_row, titulo, counter, total_valido):
        current_row = start_row
        ws.cell(row=current_row, column=1).value = titulo
        ws.cell(row=current_row, column=1).font = self.styles['title_font']
        ws.cell(row=current_row, column=1).fill = self.styles['summary_fill']
        ws.cell(row=current_row, column=1).alignment = self.styles['left_align']
        current_row += 1

        ws.cell(row=current_row, column=1).value = titulo.split()[-1].capitalize()  # Garagem ou Líder
        ws.cell(row=current_row, column=2).value = "Quantidade de Registros"
        ws.cell(row=current_row, column=3).value = "% do Total"
        for col in range(1, 4):
            cell = ws.cell(row=current_row, column=col)
            cell.font = self.styles['bold_font']
            cell.fill = self.styles['header_fill']
            cell.border = self.styles['thin_border']
            cell.alignment = self.styles['center_align']
        current_row += 1

        for item, count in counter.most_common():
            percentual = (count / total_valido * 100) if total_valido > 0 else 0
            ws.cell(row=current_row, column=1).value = item
            ws.cell(row=current_row, column=2).value = count
            ws.cell(row=current_row, column=3).value = f"{percentual:.1f}%"
            for col in range(1, 4):
                cell = ws.cell(row=current_row, column=col)
                cell.border = self.styles['thin_border']
                cell.alignment = self.styles['left_align'] if col == 1 else self.styles['center_align']
            current_row += 1

    def _add_top_section(self, ws, start_row, titulo, counter, total_valido):
        current_row = start_row
        ws.cell(row=current_row, column=1).value = titulo
        ws.cell(row=current_row, column=1).font = self.styles['title_font']
        ws.cell(row=current_row, column=1).fill = self.styles['pa_fill']
        ws.cell(row=current_row, column=1).alignment = self.styles['left_align']
        current_row += 1

        ws.cell(row=current_row, column=1).value = "Ranking"
        ws.cell(row=current_row, column=2).value = self.service.campo_especifico.capitalize()
        ws.cell(row=current_row, column=3).value = "Quantidade"
        ws.cell(row=current_row, column=4).value = "% do Total"
        for col in range(1, 5):
            cell = ws.cell(row=current_row, column=col)
            cell.font = self.styles['bold_font']
            cell.fill = self.styles['header_fill']
            cell.border = self.styles['thin_border']
            cell.alignment = self.styles['center_align']
        current_row += 1

        ranking = 1
        for item, count in counter.most_common(self.TOP_N):
            percentual = (count / total_valido * 100) if total_valido > 0 else 0
            ws.cell(row=current_row, column=1).value = f"{ranking}º"
            ws.cell(row=current_row, column=2).value = item
            ws.cell(row=current_row, column=3).value = count
            ws.cell(row=current_row, column=4).value = f"{percentual:.1f}%"
            for col in range(1, 5):
                cell = ws.cell(row=current_row, column=col)
                cell.border = self.styles['thin_border']
                cell.alignment = self.styles['center_align'] if col != 2 else self.styles['left_align']
            ranking += 1
            current_row += 1

    def _create_analise_sheet(self, wb, nome_arquivo):
        ws_analise = wb.create_sheet("Análise de Turnos")
        ws_analise.merge_cells('A1:I1')
        ws_analise['A1'] = "ANÁLISE DE TURNOS"
        ws_analise['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        ws_analise['A1'].alignment = self.styles['center_align']
        ws_analise['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

        caminho_base = None
        graficos_gerados = []
        grafico_comparacao = None
        analises_comparacao = []
        
        if self.processed_data['dados_por_dia_turno']:
            caminho_base = os.path.join(settings.MEDIA_ROOT, 
                                        nome_arquivo.replace('.xlsx', '_grafico.png'))
            
            graficos_gerados = gerar_grafico_turnos_por_mes(self.processed_data['dados_por_dia_turno'], self.inicio_dt, self.final_dt, caminho_base)
            
            caminho_comparacao = caminho_base.replace('_grafico.png', '_comparacao.png')
            resultado = gerar_grafico_comparacao_meses(self.processed_data['dados_por_dia_turno'], caminho_comparacao)
            if resultado:
                grafico_comparacao, analises_comparacao = resultado
            
            # Adicionar gráficos mensais
            linha_atual = 5
            for idx, grafico_path in enumerate(graficos_gerados):
                if grafico_path and os.path.exists(grafico_path):
                    try:
                        img = XLImage(grafico_path)
                        img.width = 1400
                        img.height = 600
                        ws_analise.add_image(img, f'A{linha_atual}')
                        linha_atual += 33
                    except Exception as e:
                        print(f"Erro ao adicionar imagem ao Excel: {e}")
            
            if grafico_comparacao and os.path.exists(grafico_comparacao):
                try:
                    img_comp = XLImage(grafico_comparacao)
                    img_comp.width = 1400
                    img_comp.height = 600
                    ws_analise.add_image(img_comp, f'A{linha_atual}')
                    linha_atual += 32
                    
                    if analises_comparacao:
                        ws_analise.cell(row=linha_atual, column=1).value = "📊 ANÁLISE DE VARIAÇÕES MENSAIS"
                        ws_analise.cell(row=linha_atual, column=1).font = Font(bold=True, size=13, color="1F4E78")
                        ws_analise.cell(row=linha_atual, column=1).fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
                        linha_atual += 1
                        
                        for analise in analises_comparacao:
                            ws_analise.cell(row=linha_atual, column=1).value = analise
                            ws_analise.cell(row=linha_atual, column=1).font = Font(size=11)
                            linha_atual += 1
                        
                        linha_atual += 2
                    
                    # Resumo Geral por Turno
                    ws_analise.cell(row=linha_atual, column=1).value = "📊 RESUMO GERAL POR TURNO"
                    ws_analise.cell(row=linha_atual, column=1).font = self.styles['title_font']
                    ws_analise.cell(row=linha_atual, column=1).fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                    linha_atual += 1
                    
                    totais_turno_resumo = {"Manhã": 0, "Tarde": 0, "Noite": 0}
                    for pa_data in self.processed_data['dados_por_pa'].values():
                        for data_turnos in pa_data.values():
                            for turno, qtd in data_turnos.items():
                                totais_turno_resumo[turno] += qtd
                    
                    total_geral_resumo = sum(totais_turno_resumo.values())
                    
                    ws_analise.cell(row=linha_atual, column=1).value = "Turno"
                    ws_analise.cell(row=linha_atual, column=2).value = "Quantidade"
                    ws_analise.cell(row=linha_atual, column=3).value = "% do Total"
                    for col in range(1, 4):
                        cell = ws_analise.cell(row=linha_atual, column=col)
                        cell.font = self.styles['bold_font']
                        cell.fill = self.styles['header_fill']
                        cell.border = self.styles['thin_border']
                        cell.alignment = self.styles['center_align']
                    linha_atual += 1
                    
                    for turno, emoji in [("Manhã", "🌅"), ("Tarde", "🌞"), ("Noite", "🌙")]:
                        ws_analise.cell(row=linha_atual, column=1).value = f"{emoji} {turno}"
                        ws_analise.cell(row=linha_atual, column=2).value = totais_turno_resumo[turno]
                        ws_analise.cell(row=linha_atual, column=3).value = f"{(totais_turno_resumo[turno]/total_geral_resumo*100):.1f}%" if total_geral_resumo > 0 else "0%"
                        for col in range(1, 4):
                            cell = ws_analise.cell(row=linha_atual, column=col)
                            cell.border = self.styles['thin_border']
                            cell.alignment = self.styles['left_align'] if col == 1 else self.styles['center_align']
                        linha_atual += 1
                    
                    ws_analise.cell(row=linha_atual, column=1).value = "📈 Total"
                    ws_analise.cell(row=linha_atual, column=2).value = total_geral_resumo
                    ws_analise.cell(row=linha_atual, column=3).value = "100%"
                    for col in range(1, 4):
                        cell = ws_analise.cell(row=linha_atual, column=col)
                        cell.border = self.styles['thin_border']
                        cell.font = self.styles['bold_font']
                        cell.fill = PatternFill(start_color="FFF4CC", end_color="FFF4CC", fill_type="solid")
                        cell.alignment = self.styles['left_align'] if col == 1 else self.styles['center_align']
                    linha_atual += 2
                    
                    # Detalhamento por P.A e Data
                    ws_analise.cell(row=linha_atual, column=1).value = "📋 DETALHAMENTO POR P.A E DATA"
                    ws_analise.cell(row=linha_atual, column=1).font = self.styles['title_font']
                    ws_analise.cell(row=linha_atual, column=1).fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                    linha_atual += 1
                    
                    ws_analise.cell(row=linha_atual, column=1).value = "Data"
                    ws_analise.cell(row=linha_atual, column=2).value = "P.A (Garagem)"
                    ws_analise.cell(row=linha_atual, column=3).value = "Turno"
                    ws_analise.cell(row=linha_atual, column=4).value = "Quantidade de Saídas"
                    for col in range(1, 5):
                        cell = ws_analise.cell(row=linha_atual, column=col)
                        cell.font = self.styles['bold_font']
                        cell.fill = self.styles['header_fill']
                        cell.border = self.styles['thin_border']
                        cell.alignment = self.styles['center_align']
                    linha_atual += 1
                    
                    for pa in sorted(self.processed_data['dados_por_pa'].keys()):
                        for data in sorted(self.processed_data['dados_por_pa'][pa].keys()):
                            for turno in ["Manhã", "Tarde", "Noite"]:
                                qtd = self.processed_data['dados_por_pa'][pa][data].get(turno, 0)
                                if qtd > 0:
                                    ws_analise.cell(row=linha_atual, column=1).value = data.strftime('%d/%m/%Y')
                                    ws_analise.cell(row=linha_atual, column=2).value = pa
                                    ws_analise.cell(row=linha_atual, column=3).value = turno
                                    ws_analise.cell(row=linha_atual, column=4).value = qtd
                                    for col in range(1, 5):
                                        cell = ws_analise.cell(row=linha_atual, column=col)
                                        cell.border = self.styles['thin_border']
                                        cell.alignment = self.styles['center_align']
                                    linha_atual += 1
                
                except Exception as e:
                    print(f"Erro ao adicionar gráfico de comparação: {e}")
        
        for col in ws_analise.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            adjusted_width = min(max_length + 3, 50)
            ws_analise.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

        return ws_analise

class RemocaoReportGenerator(BaseReportGenerator):
    def __init__(self, service):
        super().__init__(service)

# Para escalabilidade, subclasses vazias ou com overrides específicos se necessário
class DomiciliarReportGenerator(BaseReportGenerator):
    def __init__(self, service):
        super().__init__(service)

class SeletivaReportGenerator(BaseReportGenerator):
    def __init__(self, service):
        super().__init__(service)